import io
import zipfile
import openpyxl
from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from typing import List
from app.services.sql_converter import (load_mapping_from_workbook, load_date_clause_from_workbook,load_prompt_definitions, convert_sql_to_mquery)
from app.services.llm_mquery_validator import validate_mquery_with_llm
from app.schemas.stage3 import ConversionResponse, LLMFixResponse, LLMFixResult

router = APIRouter()


@router.post("/convert", response_model=ConversionResponse)
async def convert_sql_files(
    sql_files: List[UploadFile] = File(...),
    mapping_file: UploadFile = File(...),
    prompt_file: UploadFile = File(None),
):
    if not mapping_file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="mapping_file must be .xlsx")

    mapping_wb = openpyxl.load_workbook(io.BytesIO(await mapping_file.read()))
    table_mapping, column_mapping = load_mapping_from_workbook(mapping_wb)

    date_clause = ""
    prompt_definitions = {}
    if prompt_file:
        prompt_wb = openpyxl.load_workbook(io.BytesIO(await prompt_file.read()))
        date_clause = load_date_clause_from_workbook(prompt_wb)
        prompt_definitions = load_prompt_definitions(prompt_wb)


    results = []
    for sql_file in sql_files:
        if not sql_file.filename.lower().endswith(".sql"):
            continue
        raw = await sql_file.read()
        _, meta = convert_sql_to_mquery(raw, sql_file.filename, table_mapping, column_mapping, date_clause)
        results.append(meta)

    return ConversionResponse(total_files=len(results), results=results)


@router.post("/convert/download")
async def convert_and_download(
    sql_files: List[UploadFile] = File(...),
    mapping_file: UploadFile = File(...),
    prompt_file: UploadFile = File(None),
):
    mapping_wb = openpyxl.load_workbook(io.BytesIO(await mapping_file.read()))
    table_mapping, column_mapping = load_mapping_from_workbook(mapping_wb)

    date_clause = ""
    prompt_definitions = {}
    if prompt_file:
        prompt_wb = openpyxl.load_workbook(io.BytesIO(await prompt_file.read()))
        date_clause = load_date_clause_from_workbook(prompt_wb)
        prompt_definitions = load_prompt_definitions(prompt_wb)

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for sql_file in sql_files:
            if not sql_file.filename.lower().endswith(".sql"):
                continue
            raw = await sql_file.read()
            output, meta = convert_sql_to_mquery(raw, sql_file.filename, table_mapping, column_mapping, date_clause, prompt_definitions)
            zf.writestr(meta["output_filename"], output)

    zip_buf.seek(0)
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=mquery_outputs.zip"}
    )


@router.post("/convert-and-fix", response_model=LLMFixResponse)
async def convert_and_fix(
    sql_files: List[UploadFile] = File(...),
    mapping_file: UploadFile = File(...),
    prompt_file: UploadFile = File(None),
):
    """Convert SQL → M-Query, then use Databricks LLM + EXPLAIN validation loop to fix the SQL."""
    if not mapping_file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="mapping_file must be .xlsx")

    mapping_wb = openpyxl.load_workbook(io.BytesIO(await mapping_file.read()))
    table_mapping, column_mapping = load_mapping_from_workbook(mapping_wb)

    date_clause = ""
    prompt_definitions = {}
    if prompt_file:
        prompt_wb = openpyxl.load_workbook(io.BytesIO(await prompt_file.read()))
        date_clause = load_date_clause_from_workbook(prompt_wb)
        prompt_definitions = load_prompt_definitions(prompt_wb)

    results = []
    for sql_file in sql_files:
        if not sql_file.filename.lower().endswith(".sql"):
            continue
        raw = await sql_file.read()
        mquery, meta = convert_sql_to_mquery(raw, sql_file.filename, table_mapping, column_mapping, date_clause)

        final_mquery, summary, success = validate_mquery_with_llm(mquery)

        results.append(LLMFixResult(
            filename=meta["filename"],
            status="llm_fixed" if success else "llm_failed",
            output_filename=meta["output_filename"].replace("_FINAL.txt", "_LLM_FIXED.txt"),
            list_prompts_found=meta["list_prompts_found"],
            date_prompts_found=meta["date_prompts_found"],
            llm_success=success,
            llm_summary=summary,
            fixed_mquery=final_mquery,
        ))

    return LLMFixResponse(total_files=len(results), results=results)
@router.post("/validate-mquery", response_model=LLMFixResponse)
async def validate_mquery(
    mquery_files: List[UploadFile] = File(...),
):
    """Accept already-generated M-Query .txt files, extract SQL, fix with LLM, return fixed M-Query."""
    results = []
    for mquery_file in mquery_files:
        if not mquery_file.filename.lower().endswith(".txt"):
            continue
        mquery = (await mquery_file.read()).decode("utf-8", errors="replace")

        final_mquery, summary, success = validate_mquery_with_llm(mquery)

        results.append(LLMFixResult(
            filename=mquery_file.filename,
            status="llm_fixed" if success else "llm_failed",
            output_filename=mquery_file.filename.replace(".txt", "_AI_FIXED.txt"),
            list_prompts_found=0,
            date_prompts_found=0,
            llm_success=success,
            llm_summary=summary,
            fixed_mquery=final_mquery,
        ))

    return LLMFixResponse(total_files=len(results), results=results)
