import os
import re
import openpyxl
from typing import List, Tuple, Dict, Optional
from dataclasses import dataclass
from dotenv import load_dotenv
from app.core.logging import get_logger

load_dotenv()
logger = get_logger(__name__)

DATABRICKS_HOST      = os.getenv("DATABRICKS_HOST", "default-host.databricks.net")
DATABRICKS_WAREHOUSE = os.getenv("DATABRICKS_HTTP_PATH", "/sql/1.0/warehouses/default")


# ─────────────────────────────────────────────
# DATA CLASSES
# ─────────────────────────────────────────────

@dataclass
class PromptDefinition:
    prompt_name: str
    prompt_type: str   # list | single | date | like
    operator:    str   # in | = | <> | > | < | >= | <= | like | between
    data_type:   str   # string | integer | decimal | date | datetime
    allow_multi: bool
    template:    str   # list | single | date | like


@dataclass
class ExtractedPrompt:
    placeholder:    str   # __PSITEID_IN__ etc.
    prompt_name:    str
    prompt_type:    str   # promptmany | prompt
    operator:       str   # in | = | like | between | > | < | >= | <= | <>
    column_expr:    str   # LHS from SQL e.g. "ae.UnitID" or "right(ae.Patron_ID,4)"
    start_prompt:   str
    end_prompt:     str
    is_date_between: bool = False
    # OR-group: list of (column_expr) that share the same prompt via OR
    or_columns:     List[str] = None

    def __post_init__(self):
        if self.or_columns is None:
            self.or_columns = []


# ─────────────────────────────────────────────
# WORKBOOK LOADERS
# ─────────────────────────────────────────────

def load_mapping_from_workbook(wb: openpyxl.Workbook) -> Tuple[List, Dict]:
    ws      = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1)) if c.value is not None]
    idx     = {h: i for i, h in enumerate(headers)}
    table_mapping, column_mapping = [], {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            req = ['Cognos Database','Cognos Schema','Cognos Table',
                   'Databricks Database','Databricks Schema','Databricks Table']
            if all(k in idx for k in req):
                cp = [row[idx['Cognos Database']], row[idx['Cognos Schema']], row[idx['Cognos Table']]]
                dp = [row[idx['Databricks Database']], row[idx['Databricks Schema']], row[idx['Databricks Table']]]
                if all(p is not None for p in cp + dp):
                    table_mapping.append((f"{cp[0]}.{cp[1]}.{cp[2]}", f"{dp[0]}.{dp[1]}.{dp[2]}"))
            cog_col = row[idx["Cognos Columns"]]    if "Cognos Columns"    in idx else None
            db_col  = row[idx["Databricks Columns"]] if "Databricks Columns" in idx else None
            if cog_col and db_col:
                column_mapping[str(cog_col).lower().strip()] = str(db_col).strip()
        except Exception as e:
            logger.warning(f"Skipping malformed mapping row: {e}")
    return table_mapping, column_mapping


def load_prompt_definitions(wb: openpyxl.Workbook) -> Dict[str, PromptDefinition]:
    definitions: Dict[str, PromptDefinition] = {}
    if "PromptDefinitions" not in wb.sheetnames:
        return definitions
    ws      = wb["PromptDefinitions"]
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
    idx     = {h: i for i, h in enumerate(headers) if h}

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            def g(col):
                return str(row[idx[col]]).strip() if col in idx and row[idx[col]] is not None else ""
            name = g("PromptName")
            if not name:
                continue
            allow_raw = g("AllowMultiSelect").lower()
            definitions[name.lower()] = PromptDefinition(
                prompt_name = name,
                prompt_type = g("PromptType").lower()  or "single",
                operator    = g("Operator").lower()    or "in",
                data_type   = g("DataType").lower()    or "string",
                allow_multi = allow_raw in ("true", "yes", "1"),
                template    = g("Template").lower()    or "single",
            )
        except Exception as e:
            logger.warning(f"Skipping PromptDefinitions row: {e}")
    return definitions


def load_date_clause_from_workbook(wb: openpyxl.Workbook) -> str:
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell:
                text = str(cell)
                if "if (" in text.lower() and "date" in text.lower():
                    return text.strip()
    return ""


# ─────────────────────────────────────────────
# SQL CLEANING HELPERS
# ─────────────────────────────────────────────

def remove_double_quotes(sql: str) -> str:
    sql = re.sub(r'"(\w+)"\.\"(\w+)"', r'\1.\2', sql)
    return re.sub(r'"(\w+)"', r'\1', sql)


def fix_databricks_sql(sql: str) -> str:
    sql = re.sub(r"\[(.*?)\]", r"\1", sql)
    sql = re.sub(r"\(\s*nolock\s*\)", "", sql, flags=re.IGNORECASE)
    # Replace SPACE(n) with a literal string of n spaces — Databricks doesn't have SPACE()
    def replace_space_fn(m: re.Match) -> str:
        try:
            n = int(m.group(1).strip())
            return f"'{' ' * n}'"
        except ValueError:
            return "' '"
    sql = re.sub(r'\bSPACE\s*\(\s*(\d+)\s*\)', replace_space_fn, sql, flags=re.IGNORECASE)
    def fix_legacy_concat(m):
        alias   = m.group(1).strip()
        content = m.group(2).strip()
        if content.startswith('(') and content.endswith(')') and content.count('(') == content.count(')'):
            content = content[1:-1].strip()
        args = re.sub(r"\s*\+\s*", ", ", content)
        args = re.sub(r"\bvarchar\s*\(\d+\)", "STRING", args, flags=re.IGNORECASE)
        return f"CONCAT({args}) AS {alias}"

    sql = re.sub(r"\b([\w_]+)\s*=\s*([^\n,]+(?:\+)[^\n,]+)", fix_legacy_concat, sql, flags=re.IGNORECASE)

    def fix_subquery_alias(m):
        return f"{m.group(2).strip()} AS {m.group(1)}"
    sql = re.sub(r"\b(\w+)\s*=\s*(\(\s*SELECT[\s\S]*?\))", fix_subquery_alias, sql, flags=re.IGNORECASE)

    # NEW: fix plain col = expr aliases (no + and no subquery) → expr AS col
    # Only matches inside SELECT lists: alias = table.column or alias = function(...)
    def fix_plain_alias(m):
        alias = m.group(1).strip()
        expr  = m.group(2).strip()
        return f"{expr} AS {alias}"

    _protected_line = re.compile(
        r'^\s*(?:ON\b|WHERE\b|HAVING\b|(?:INNER|LEFT|RIGHT|FULL|CROSS)?\s*JOIN\b)',
        re.IGNORECASE
    )
    fixed_lines = []
    for line in sql.splitlines():
        if _protected_line.match(line):
            fixed_lines.append(line)
        else:
            fixed_lines.append(re.sub(
                r'\b([A-Za-z_]\w*)\s*=\s*((?!null\b|\'|"|CASE\b)[A-Za-z_][\w.]*(?:\s*\([^()]*\))?)',
                fix_plain_alias,
                line,
                flags=re.IGNORECASE
            ))
    sql = "\n".join(fixed_lines)

    return sql

def _remap_bare_column_outside_joins(sql: str, cog_col: str, db_col: str) -> str:
    """
    Replace bare `cog_col` with `db_col` only on lines that are NOT part of
    a FROM / JOIN clause, to avoid replacing table aliases.
    """
    join_line_pat = re.compile(
        r'^\s*(?:FROM|ON|(?:INNER|LEFT|RIGHT|FULL|CROSS)?\s*JOIN)\b',
        re.IGNORECASE
    )
    col_pat = re.compile(rf'(?<!\.)\b{re.escape(cog_col)}\b', re.IGNORECASE)

    result_lines = []
    for line in sql.splitlines():
        if join_line_pat.match(line):
            result_lines.append(line)  # leave FROM/JOIN lines untouched for bare replacement
        else:
            result_lines.append(col_pat.sub(db_col, line))
    return "\n".join(result_lines)

def apply_mapping(sql: str, table_mapping: List, column_mapping: Dict) -> str:
    # Table mapping — safe globally (full three-part names won't collide)
    for cog, db in sorted(table_mapping, key=lambda x: len(x[0]), reverse=True):
        sql = re.sub(re.escape(re.sub(r"[\[\]]", "", cog)), db, sql, flags=re.IGNORECASE)

    if not column_mapping:
        return sql

    for cog_col, db_col in sorted(column_mapping.items(), key=lambda x: len(x[0]), reverse=True):
        # 1. Always remap alias.column → alias.new_column (safe everywhere including JOIN ON)
        sql = re.sub(
            rf'(\b\w+\.){re.escape(cog_col)}\b',
            rf'\1{db_col}',
            sql, flags=re.IGNORECASE
        )
        # 2. Bare column replacement — ONLY in SELECT, WHERE, HAVING, GROUP BY, ORDER BY
        #    Never in FROM/JOIN lines where bare tokens are table aliases not column names
        sql = _remap_bare_column_outside_joins(sql, cog_col, db_col)

    return sql

def ensure_where_1_1(sql: str) -> str:
    """
    For every WHERE clause that has no static condition (i.e. is immediately
    followed by a placeholder token or end-of-clause), inject 1=1 so that
    dynamic AND filters are always safe.
    Also inserts WHERE 1=1 if no WHERE exists at all.
    """
    # If no WHERE at all, insert one before GROUP BY / HAVING / ORDER BY
    if not re.search(r'\bWHERE\b', sql, re.IGNORECASE):
        anchor = "WHERE 1=1"
        pat = re.compile(r'\b(GROUP\s+BY|HAVING|ORDER\s+BY)\b', re.IGNORECASE)
        m = pat.search(sql)
        if m:
            return sql[:m.start()].rstrip() + f"\n{anchor}\n" + sql[m.start():]
        return sql.rstrip() + f"\n{anchor}\n"

    # Inject 1=1 after WHERE when the very next non-space token is a placeholder
    # (__SOMETHING__) or end of string — meaning no static condition exists
    sql = re.sub(
        r'\bWHERE\s+(?=__[A-Z0-9_]+__|$)',
        'WHERE 1=1\n',
        sql,
        flags=re.IGNORECASE,
    )
    return sql


def inject_filters_before_clause(sql: str, filter_fragment: str) -> str:
    """Insert filter_fragment before GROUP BY / HAVING / ORDER BY, or append."""
    pat = re.compile(r'\b(GROUP\s+BY|HAVING|ORDER\s+BY)\b', re.IGNORECASE)
    m   = pat.search(sql)
    if m:
        return sql[:m.start()].rstrip() + "\n" + filter_fragment + "\n" + sql[m.start():]
    return sql.rstrip() + "\n" + filter_fragment


# ─────────────────────────────────────────────
# PROMPT EXTRACTION
# ─────────────────────────────────────────────

_PROMPT_EXPR = r'#\s*prompt(?:many)?\s*\(\s*\'([^\']+)\'\s*\)\s*#'

# BETWEEN with optional CAST(...) / DATE(...) wrappers around each prompt
_BETWEEN_PAT = re.compile(
    r'([\w.]+(?:\s*\(.*?\))?)\s+BETWEEN\s+'
    r'(?:(?:CAST|DATE)\s*\(\s*)?' + _PROMPT_EXPR + r'(?:\s+AS\s+\w+\s*)?\)?\s*'
    r'AND\s+'
    r'(?:(?:CAST|DATE)\s*\(\s*)?' + _PROMPT_EXPR + r'(?:\s+AS\s+\w+\s*)?\)?',
    re.IGNORECASE | re.DOTALL,
)

# col IN (#prompt(...)# or #promptmany(...)#)
_IN_PAT = re.compile(
    r'([\w.]+(?:\s*\(.*?\))?)\s+IN\s*\(\s*' + _PROMPT_EXPR + r'\s*\)',
    re.IGNORECASE,
)
# col IN #prompt(...)#  — without parentheses (non-standard Cognos syntax)
_IN_NO_PARENS_PAT = re.compile(
    r'([\w.]+(?:\s*\(.*?\))?)\s+IN\s+' + _PROMPT_EXPR,
    re.IGNORECASE,
)
# col OP #prompt(...)#   (=, <>, !=, >=, <=, >, <, LIKE)
_OP_PAT = re.compile(
    r'([\w.]+(?:\s*\(.*?\))?)\s*(LIKE|<>|!=|>=|<=|>|<|=)\s*' + _PROMPT_EXPR,
    re.IGNORECASE,
)


def _make_placeholder(prompt_name: str, suffix: str = "") -> str:
    safe = re.sub(r"[^a-zA-Z0-9]", "_", prompt_name).upper()
    return f"__{safe}{suffix}__"


def _var_name(prompt_name: str) -> str:
    return re.sub(r"[^a-zA-Z0-9]", "", prompt_name)


def extract_prompts(sql: str) -> Tuple[str, List[ExtractedPrompt]]:
    """
    Replace every Cognos #prompt# / #promptmany# expression with a unique
    placeholder token.  Returns (modified_sql, list[ExtractedPrompt]).

    OR-groups: when the same prompt name appears multiple times in an OR block
    (e.g. col1 IN (#prompt('X')#) OR col2 IN (#prompt('X')#)) we merge them
    into a single ExtractedPrompt with or_columns populated, so one M variable
    covers the whole OR expression.
    """
    prompts:          List[ExtractedPrompt] = []
    used_ph:          set                   = set()
    # prompt_name.lower() → ExtractedPrompt (for OR-group merging)
    by_name:          Dict[str, ExtractedPrompt] = {}

    def unique_ph(base: str) -> str:
        ph, i = base, 1
        while ph in used_ph:
            ph = f"{base}_{i}"; i += 1
        used_ph.add(ph)
        return ph

    # ── 1. BETWEEN ────────────────────────────────────────────────────────
    def replace_between(m: re.Match) -> str:
        col_expr   = m.group(1).strip()
        start_name = m.group(2).strip()
        end_name   = m.group(3).strip()
        ph = unique_ph(_make_placeholder(start_name, "_DATE"))
        ep = ExtractedPrompt(
            placeholder=ph, prompt_name=start_name, prompt_type="prompt",
            operator="between", column_expr=col_expr,
            start_prompt=start_name, end_prompt=end_name, is_date_between=True,
        )
        prompts.append(ep)
        by_name[start_name.lower()] = ep
        return ph

    sql = _BETWEEN_PAT.sub(replace_between, sql)

    # ── 2. IN ─────────────────────────────────────────────────────────────
    def replace_in(m: re.Match) -> str:
        col_expr    = m.group(1).strip()
        prompt_name = m.group(2).strip()
        key         = prompt_name.lower()
        ptype       = "promptmany" if re.search(r'promptmany', m.group(0), re.I) else "prompt"

        if key in by_name and by_name[key].operator == "in":
            # Same prompt used again in an OR — add this column to the OR-group
            existing = by_name[key]
            if col_expr not in existing.or_columns and col_expr != existing.column_expr:
                existing.or_columns.append(col_expr)
            return existing.placeholder
        else:
            ph = unique_ph(_make_placeholder(prompt_name, "_IN"))
            ep = ExtractedPrompt(
                placeholder=ph, prompt_name=prompt_name, prompt_type=ptype,
                operator="in", column_expr=col_expr,
                start_prompt=prompt_name, end_prompt=prompt_name,
            )
            prompts.append(ep)
            by_name[key] = ep
            return ph

    sql = _IN_PAT.sub(replace_in, sql)
    sql = _IN_NO_PARENS_PAT.sub(replace_in, sql)
    # ── 3. Generic operator ───────────────────────────────────────────────
    def replace_op(m: re.Match) -> str:
        col_expr    = m.group(1).strip()
        operator    = m.group(2).strip().lower()
        prompt_name = m.group(3).strip()
        ph = unique_ph(_make_placeholder(prompt_name, f"_{operator.upper()}"))
        ep = ExtractedPrompt(
            placeholder=ph, prompt_name=prompt_name, prompt_type="prompt",
            operator=operator, column_expr=col_expr,
            start_prompt=prompt_name, end_prompt=prompt_name,
        )
        prompts.append(ep)
        by_name[prompt_name.lower()] = ep
        return ph

    sql = _OP_PAT.sub(replace_op, sql)

    # ── 4. Bare fallback ──────────────────────────────────────────────────
    def replace_bare(m: re.Match) -> str:
        prompt_name = m.group(1).strip()
        ph = unique_ph(_make_placeholder(prompt_name, "_BARE"))
        ep = ExtractedPrompt(
            placeholder=ph, prompt_name=prompt_name, prompt_type="prompt",
            operator="=", column_expr="",
            start_prompt=prompt_name, end_prompt=prompt_name,
        )
        prompts.append(ep)
        return ph

    sql = re.sub(_PROMPT_EXPR, replace_bare, sql, flags=re.IGNORECASE)

    return sql, prompts


# ─────────────────────────────────────────────
# M-QUERY VARIABLE BUILDERS
# ─────────────────────────────────────────────

def _item_fmt(data_type: str) -> str:
    """M expression to format one list item for SQL embedding."""
    if data_type in ("string", "date", "datetime"):
        return "'\" & Text.From(_) & \"'"
    return "\" & Text.From(_) & \""   # numeric — no quotes


def _scalar_fmt(var_ref: str, data_type: str) -> str:
    """M expression to format a scalar value for SQL embedding."""
    if data_type in ("string", "date", "datetime"):
        return f"\"'\" & Text.From({var_ref}) & \"'\""
    return f"Text.From({var_ref})"


def _build_list_template(ep: ExtractedPrompt, defn: Optional[PromptDefinition]) -> Tuple[str, str]:
    pname     = ep.prompt_name
    vn        = _var_name(pname)
    data_type = defn.data_type if defn else "string"

    if data_type in ("string", "date", "datetime"):
        item_expr = "\"'\" & Text.From(_) & \"'\""
    else:
        item_expr = "Text.From(_)"

    # Build the IN condition — may be a single column or an OR-group
    all_cols = [ep.column_expr] + ep.or_columns
    if len(all_cols) == 1:
        in_condition = f'" AND {all_cols[0]} IN (" & combinedlist_{vn} & ")"'
    else:
        parts = [f'{c} IN (" & combinedlist_{vn} & ")' for c in all_cols]
        or_body = "\n                OR ".join(parts)
        in_condition = f'" AND (\n                {or_body}\n            )"'

    block = f"""
    combinedlist_{vn} =
        if List.Count({pname}) = 0 then ""
        else Text.Combine(List.Transform({pname}, each {item_expr}), ","),

    where{vn} =
        if combinedlist_{vn} <> ""
        then {in_condition}
        else "",
"""
    return block, f"where{vn}"


def _build_single_template(ep: ExtractedPrompt, defn: Optional[PromptDefinition]) -> Tuple[str, str]:
    pname     = ep.prompt_name
    vn        = _var_name(pname)
    data_type = defn.data_type if defn else "string"
    col       = ep.column_expr
    op        = (defn.operator if defn else ep.operator).upper()
    val_expr  = _scalar_fmt(pname, data_type)

    block = f"""
    where{vn} =
        if {pname} = null then ""
        else " AND {col} {op} " & {val_expr},
"""
    return block, f"where{vn}"


def _build_like_template(ep: ExtractedPrompt, defn: Optional[PromptDefinition]) -> Tuple[str, str]:
    pname = ep.prompt_name
    vn    = _var_name(pname)
    col   = ep.column_expr

    block = f"""
    where{vn} =
        if {pname} = null then ""
        else " AND {col} LIKE '%" & Text.From({pname}) & "%'",
"""
    return block, f"where{vn}"


def _build_date_template(ep: ExtractedPrompt, defn: Optional[PromptDefinition]) -> Tuple[str, str]:
    start     = ep.start_prompt
    end       = ep.end_prompt
    col       = ep.column_expr
    data_type = defn.data_type if defn else "date"
    fmt       = "yyyy-MM-dd HH:mm:ss" if data_type == "datetime" else "yyyy-MM-dd"
    fn        = "DateTime.ToText" if data_type == "datetime" else "Date.ToText"

    block = f"""
    dateClause =
        if {start} <> null and {end} <> null then
            \" AND {col} BETWEEN '\" &
            {fn}({start}, \"{fmt}\") &
            \"' AND '\" &
            {fn}({end}, \"{fmt}\") &
            \"'\"
        else if {start} <> null then
            \" AND {col} >= '\" & {fn}({start}, \"{fmt}\") & \"'\"
        else if {end} <> null then
            \" AND {col} <= '\" & {fn}({end}, \"{fmt}\") & \"'\"
        else \"\",
"""
    return block, "dateClause"



def build_prompt_variables(
    prompts:     List[ExtractedPrompt],
    definitions: Dict[str, PromptDefinition],
) -> Tuple[List[str], Dict[str, str]]:
    """
    Returns:
        m_blocks  — M variable definition strings (one per unique prompt)
        ph_to_var — placeholder → M variable name
    """
    m_blocks:  List[str]       = []
    ph_to_var: Dict[str, str]  = {}
    seen_vars: set              = set()

    for ep in prompts:
        defn     = definitions.get(ep.prompt_name.lower())
        template = defn.template if defn else (
            "date"   if ep.is_date_between  else
            "list"   if ep.operator == "in" else
            "like"   if ep.operator == "like" else
            "single"
        )

        if ep.is_date_between:
            block, var_name = _build_date_template(ep, defn)
        elif template == "list":
            block, var_name = _build_list_template(ep, defn)
        elif template == "like":
            block, var_name = _build_like_template(ep, defn)
        else:
            block, var_name = _build_single_template(ep, defn)

        if var_name not in seen_vars:
            m_blocks.append(block)
            seen_vars.add(var_name)

        ph_to_var[ep.placeholder] = var_name

    return m_blocks, ph_to_var


# ─────────────────────────────────────────────
# M-QUERY ASSEMBLY  (Issue 1 fix: segment-based concatenation)
# ─────────────────────────────────────────────

_PLACEHOLDER_RE = re.compile(r'(__[A-Z0-9_]+__)')


def _split_sql_on_placeholders(sql: str, ph_to_var: Dict[str, str]) -> List[Tuple[str, str]]:
    """
    Split sql into alternating (literal_segment, var_name) pairs.
    The last tuple always has var_name = "".
    Example:
        "SELECT * FROM t WHERE " + PH1 + " AND " + PH2 + " ORDER BY x"
        → [("SELECT * FROM t WHERE ", "whereA"),
           (" AND ", "whereB"),
           (" ORDER BY x", "")]
    """
    parts: List[Tuple[str, str]] = []
    pos = 0
    for m in _PLACEHOLDER_RE.finditer(sql):
        literal = sql[pos:m.start()]
        ph      = m.group(1)
        var     = ph_to_var.get(ph, "")
        parts.append((literal, var))
        pos = m.end()
    parts.append((sql[pos:], ""))
    return parts


def _escape_m_string(s: str) -> str:
    """Escape a literal SQL fragment for embedding in an M double-quoted string."""
    return s.replace('"', '""')

def validate_sql_aliases(sql: str) -> List[str]:
    """
    Check that every alias used in SELECT exists in FROM/JOIN.
    Returns list of warning strings. Empty = clean.
    """
    warnings = []

    # Extract all defined aliases from FROM / JOIN clauses
    # Matches: table_or_subquery [AS] alias
    defined = set(re.findall(
        r'(?:FROM|JOIN)\s+[\w.]+\s+(?:AS\s+)?(\w+)',
        sql, re.IGNORECASE
    ))
    # Also catch subquery aliases: ) alias or ) AS alias
    defined |= set(re.findall(r'\)\s+(?:AS\s+)?(\w+)', sql, re.IGNORECASE))

    # Extract alias prefixes used in SELECT columns: alias.column
    used = set(re.findall(r'\b(\w+)\.\w+', sql, re.IGNORECASE))

    # SQL keywords and common non-alias prefixes to ignore
    keywords = {
        'select','from','where','join','on','and','or','not','in','as','by',
        'group','order','having','left','right','inner','outer','full','cross',
        'silver_prod','gold_prod','bronze','dbo','operations','accounting',  # catalog/schema names
    }

    phantom = used - defined - keywords
    for alias in sorted(phantom):
        warnings.append(f"Alias '{alias}' used in SELECT but not found in FROM/JOIN clauses.")

    return warnings

def build_m_query(
    sql:                str,
    prompts:            List[ExtractedPrompt],
    definitions:        Dict[str, PromptDefinition],
    legacy_date_clause: str = "",
) -> str:
    m_blocks, ph_to_var = build_prompt_variables(prompts, definitions)

    # Legacy date clause fallback
    if legacy_date_clause and not any(ep.is_date_between for ep in prompts):
        m_blocks.append(f"\n    {legacy_date_clause},")
        ph_to_var["__LEGACY_DATE__"] = "dateClause"
        sql = inject_filters_before_clause(sql, "__LEGACY_DATE__")

    # Ensure WHERE 1=1 exists so AND filters always have an anchor
    sql = ensure_where_1_1(sql)

    # Split SQL on placeholders and build M string concatenation
    segments = _split_sql_on_placeholders(sql, ph_to_var)

    # Build the query = ... expression as proper M string concatenation
    # Each literal segment is a quoted M string; each var is a bare M identifier
    query_lines: List[str] = []
    need_amp = False  # tracks whether the next token needs a leading &
    for literal, var_name in segments:
        escaped = _escape_m_string(literal)
        if escaped:
            prefix = "        & " if need_amp else "        "
            query_lines.append(f'{prefix}"{escaped}"')
            need_amp = True
        if var_name:
            query_lines.append(f'        & {var_name}')
            need_amp = True

    query_expr = "\n".join(query_lines)

    host         = re.sub(r'^https?://', '', DATABRICKS_HOST).rstrip('/')
    warehouse_id = DATABRICKS_WAREHOUSE.split("/")[-1]
    var_block    = "".join(m_blocks)

    return f"""let
{var_block}
    query =
{query_expr},

    Source = Databricks.Catalogs(
        "{host}",
        "/sql/1.0/warehouses/{warehouse_id}",
        [EnableAutomaticProxyDiscovery = false]
    ),

    finalsource = Value.NativeQuery(
        Source,
        query,
        null,
        [EnableFolding = false]
    )

in
    finalsource"""


# ─────────────────────────────────────────────
# MAIN ENTRY POINT
# ─────────────────────────────────────────────

def convert_sql_to_mquery(
    sql_raw:            bytes,
    filename:           str,
    table_mapping:      List,
    column_mapping:     Dict,
    date_clause:        str,
    prompt_definitions: Optional[Dict[str, PromptDefinition]] = None,
) -> Tuple[str, Dict]:
    from app.services.cognos_cleanup import read_text, strip_cognos_wrappers

    sql = read_text(sql_raw)
    sql = strip_cognos_wrappers(sql)

    sql, prompts = extract_prompts(sql)
    sql = remove_double_quotes(sql)
    sql = fix_databricks_sql(sql)
    sql = apply_mapping(sql, table_mapping, column_mapping)

    # Pre-generation validation
    sql_warnings = validate_sql_aliases(sql)

    # Unresolved placeholders check
    unresolved = re.findall(r'__[A-Z0-9_]+__', sql)
    if unresolved:
        sql_warnings.append(f"Unresolved prompt placeholders: {list(set(unresolved))}")

    # GROUP BY after WHERE check
    if re.search(r'\bGROUP\s+BY\b.*\bWHERE\b', sql, re.IGNORECASE | re.DOTALL):
        sql_warnings.append("GROUP BY appears before WHERE — clause ordering may be incorrect.")

    for w in sql_warnings:
        logger.warning(f"[SQL Validation] {filename}: {w}")

    output = build_m_query(sql, prompts, prompt_definitions or {}, date_clause)

    return output, {
        "filename":           filename,
        "status":             "success",
        "output_filename":    filename.replace(".sql", "_Mquery_FINAL.txt"),
        "list_prompts_found": sum(1 for p in prompts if p.operator == "in"),
        "date_prompts_found": sum(1 for p in prompts if p.is_date_between),
        "sql_warnings":       sql_warnings,
    }

